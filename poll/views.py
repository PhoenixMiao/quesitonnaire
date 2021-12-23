from poll.models import Questionnaire, Whitelist, Blacklist,Record,HistoryRecord,Condition,Choice
from django.http import JsonResponse
from http import HTTPStatus
from django.forms.models import model_to_dict
from Utils.wrappers import permitted_methods
from Utils.tools import request_body_serialize,request_body_serialize_init
from django.core.paginator import Paginator,EmptyPage
from student.models import Student
from Utils.tools import paginator2dict,paginator3dict,paginator4dict,paginator5dict,paginator6dict
from Utils.fileTool import upload_file
import xlwt
import io, csv, codecs
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl import load_workbook
import wx
import wx.grid


@permitted_methods(["GET"])
def polls(request, type):
    page_num = request.GET.get('p', 1)
    length = request.GET.get('l', 15)
    if type == '' or type is None:
        # 获取未归档的问卷列表
        questionnaires = Questionnaire.objects.exclude(status=-1)
        paginator = Paginator(questionnaires, length)
        try:
            paginator_page = paginator.page(page_num)
        except EmptyPage:
            return JsonResponse(status=HTTPStatus.NO_CONTENT, data={"error": "没有该页面"},
                                json_dumps_params={'ensure_ascii': False})
        ret = {'message': 'ok',
               'data': paginator4dict(paginator_page, ["id", "title", "creatorId", "oneoff", "status",'createTime','updateTime'])}
        return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})
    elif type == 'archive':
        # 获取已归档的问卷列表：
        questionnaires = Questionnaire.objects.filter(status=-1)
        paginator = Paginator(questionnaires, length)
        try:
            paginator_page = paginator.page(page_num)
        except EmptyPage:
            return JsonResponse(status=HTTPStatus.NO_CONTENT, data={"error": "没有该页面"},
                                json_dumps_params={'ensure_ascii': False})
        ret = {'message': 'ok',
               'data': paginator5dict(paginator_page,["id", "title", "creatorId", "oneoff", "status", 'createTime', 'updateTime'])}
        return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})
    else:
        return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '参数错误'},
                            json_dumps_params={'ensure_ascii': False})




@permitted_methods(["GET"])
def meta(request, pollId):
    questionnaires = Questionnaire.objects.filter(id=pollId)
    if len(questionnaires) == 0:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该问卷'},
                            json_dumps_params={'ensure_ascii': False})
    questionnaire = questionnaires[0]
    fields = ["status", "oneoff", "title", "scope", "creatorId"]
    for i in range(1, 21):
        if getattr(questionnaire, "k" + str(i)) is not None:
            fields.append("k" + str(i))
    mod = model_to_dict(questionnaire, fields=fields)
    mod['createTime'] = questionnaire.createTime.strftime("%Y-%m-%d %H:%M")
    mod['updateTime'] = questionnaire.updateTime.strftime("%Y-%m-%d %H:%M")
    ret = {'message': 'ok',
           'data': mod}
    return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["POST"])
def whitelist_search(request,pollId):
    page_num = request.GET.get('p', 1)
    length = request.GET.get('l', 15)
    vars = request_body_serialize(request)
    whiteList = Whitelist.objects.filter(xh__icontains=vars['xh'],questionnaireId=pollId)
    if(len(whiteList)==0):
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={"error": "该学生并不在白名单内"},
                            json_dumps_params={'ensure_ascii': False})
    paginator = Paginator(whiteList,length)
    try:
        paginator_page = paginator.page(page_num)
    except EmptyPage:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={"error": "没有该页面"},
                            json_dumps_params={'ensure_ascii': False})
    ret = {'message': 'ok',
           'data': paginator6dict(paginator_page,['questionnaireId','xh'])}
    return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})




@permitted_methods(["DELETE"])
def whitelist_delete(request, pollId):
    vars = request_body_serialize(request)
    if "xh" not in vars.keys():
        return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '缺少参数'},
                            json_dumps_params={'ensure_ascii': False})
    whitelist = Whitelist.objects.filter(questionnaireId=pollId, xh=vars['xh'])
    if len(whitelist) == 0:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '白名单或名单中学号不存在'},
                            json_dumps_params={'ensure_ascii': False})
    whitelist.delete()
    return JsonResponse(data={'message': 'ok'}, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["POST"])
def whitelist_import(request, pollId):
    tmpFile = request.FILES.get('whitelist')
    workbook_ = load_workbook(tmpFile)
    sheetnames = workbook_.get_sheet_names()
    sheet = workbook_.get_sheet_by_name(sheetnames[0])
    j=2
    tmp = sheet.cell(row=2,column=1).value
    while tmp:
        student = Student.objects.filter(xh=tmp)
        if (len(student) == 0):
            return JsonResponse(status=HTTPStatus.NO_CONTENT, data={"error": "没有该学生"},
                                json_dumps_params={'ensure_ascii': False})
        blacklist = Blacklist.objects.filter(xh=tmp)
        if (len(blacklist) != 0):
            return JsonResponse(status=HTTPStatus.NO_CONTENT, data={"error": "该学生已经在黑名单里了"},
                                json_dumps_params={'ensure_ascii': False})
        Whitelist.objects.create(questionnaireId=pollId, xh=tmp)
        j = j+1
        tmp = sheet.cell(row=j,column=1).value
    return JsonResponse(data={'message':'ok'}, json_dumps_params={'ensure_ascii': False})



@permitted_methods(["POST"])
def blacklist_search(request,pollId):
    page_num = request.GET.get('p', 1)
    length = request.GET.get('l', 15)
    vars = request_body_serialize(request)
    blackList = Blacklist.objects.filter(xh__icontains=vars['xh'],questionnaireId=pollId)
    if(len(blackList)==0):
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={"error": "该学生并不在黑名单内"},
                            json_dumps_params={'ensure_ascii': False})
    paginator = Paginator(blackList,length)
    try:
        paginator_page = paginator.page(page_num)
    except EmptyPage:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={"error": "没有该页面"},
                            json_dumps_params={'ensure_ascii': False})
    ret = {'message': 'ok',
           'data': paginator6dict(paginator_page,['questionnaireId','xh'])}
    return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["DELETE"])
def blacklist_delete(request, pollId):
    vars = request_body_serialize(request)
    if "xh" not in vars.keys():
        return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '缺少参数'},
                            json_dumps_params={'ensure_ascii': False})
    blacklist = Blacklist.objects.filter(questionnaireId=pollId, xh=vars['xh'])
    if len(blacklist) == 0:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '白名单或名单中学号不存在'},
                            json_dumps_params={'ensure_ascii': False})
    blacklist.delete()
    return JsonResponse(data={'message': 'ok'}, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["POST"])
def blacklist_import(request, pollId):
    # reader = csv.reader(request)
    tmpFile = request.FILES.get('blacklist')
    workbook_ = load_workbook(tmpFile)
    sheetnames = workbook_.get_sheet_names()
    sheet = workbook_.get_sheet_by_name(sheetnames[0])
    j=2
    tmp = sheet.cell(row=2,column=1).value
    while tmp:
        student = Student.objects.filter(xh=tmp)
        if (len(student) == 0):
            return JsonResponse(status=HTTPStatus.NO_CONTENT, data={"error": "没有该学生"},
                                json_dumps_params={'ensure_ascii': False})
        whitelist = Whitelist.objects.filter(xh=tmp)
        if (len(whitelist) != 0):
            return JsonResponse(status=HTTPStatus.NO_CONTENT, data={"error": "该学生已经在白名单里了"},
                                json_dumps_params={'ensure_ascii': False})
        Blacklist.objects.create(questionnaireId=pollId, xh=tmp)
        j = j+1
        tmp = sheet.cell(row=j,column=1).value
    return JsonResponse(data={'message':'ok'}, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["POST"])
def poll_create(request):
    body_list = request_body_serialize_init(request)
    status = int(body_list.get("status"))
    oneoff = int(body_list.get("oneoff"))
    title = body_list.get("title")
    scope = int(body_list.get("scope"))
    creatorId = '20130053'
    sub_body_list = list(body_list.values())[list(body_list.keys()).index("1"):]
    size = len(sub_body_list)
    for i in range(size+1,21):
        sub_body_list.append(None)
    Questionnaire.objects.create(status=status, oneoff=oneoff,title=title, scope=scope, creatorId=creatorId,
                                 k1=sub_body_list[0], k2=sub_body_list[1], k3=sub_body_list[2], k4=sub_body_list[3],
                                 k5=sub_body_list[4], k6=sub_body_list[5], k7=sub_body_list[6], k8=sub_body_list[7],
                                 k9=sub_body_list[8], k10=sub_body_list[9], k11=sub_body_list[10], k12=sub_body_list[11],
                                 k13=sub_body_list[12], k14=sub_body_list[13], k15=sub_body_list[14], k16=sub_body_list[15],
                                 k17=sub_body_list[16], k18=sub_body_list[17], k19=sub_body_list[18], k20=sub_body_list[19])
    return JsonResponse(data={'message': 'ok'}, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["POST"])
def poll_activate(request,pollId):
    questionnaires = Questionnaire.objects.filter(id=pollId)
    if len(questionnaires) == 0:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该问卷'},
                            json_dumps_params={'ensure_ascii': False})
    else:
        questionnaire = questionnaires[0]
        if questionnaire.status == 0 or questionnaire.status == 2:
           questionnaires[0].status = 1
        elif questionnaire.status == 1:
            return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '该问卷已经在发布状态'},
                            json_dumps_params={'ensure_ascii': False})
        elif questionnaire.status == -1:
            return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '该问卷已归档'},
                            json_dumps_params={'ensure_ascii': False})
        questionnaires[0].save()
        return JsonResponse(data={'message': 'ok'}, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["POST"])
def poll_pause(request,pollId):
    questionnaires = Questionnaire.objects.filter(id=pollId)
    if len(questionnaires) == 0:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该问卷'},
                            json_dumps_params={'ensure_ascii': False})
    else:
        questionnaire = questionnaires[0]
        if questionnaire.status == 1:
           questionnaires[0].status = 2
        elif questionnaire.status == 2:
            return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '该问卷已经暂停发布'},
                            json_dumps_params={'ensure_ascii': False})
        elif questionnaire.status == 0:
            return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '该问卷仍在草稿状态'},
                            json_dumps_params={'ensure_ascii': False})
        elif questionnaire.status == -1:
            return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '该问卷已归档'},
                            json_dumps_params={'ensure_ascii': False})
        questionnaires[0].save()
        return JsonResponse(data={'message': 'ok'}, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["POST"])
def record_add_judge(request):
    body_list = request_body_serialize_init(request)
    xh = body_list.get('xh')
    questionnaire_id = body_list.get('questionnaireId')
    questionnaires = Questionnaire.objects.filter(id=questionnaire_id)
    questionnaire = Questionnaire.objects.get(id=questionnaire_id)
    if len(questionnaires) == 0:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该问卷'},
                            json_dumps_params={'ensure_ascii': False})
    stu = Student.objects.filter(xh=xh)
    if len(stu)==0:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该学生'},
                            json_dumps_params={'ensure_ascii': False})
    whitelist = Whitelist.objects.filter(questionnaireId=questionnaire_id,xh=xh)
    blacklist = Blacklist.objects.filter(questionnaireId=questionnaire_id,xh=xh)
    if len(whitelist)==0:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '该学生并不在白名单内'},
                            json_dumps_params={'ensure_ascii': False})
    if len(blacklist)!=0:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '该学生并在黑名单内'},
                            json_dumps_params={'ensure_ascii': False})
    questionnaire_dic = model_to_dict(questionnaire)
    if questionnaire_dic.get("status") != 1:
        return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '该问卷并不在发布状态'},
                            json_dumps_params={'ensure_ascii': False})
    recorded = Record.objects.filter(questionnaireId=questionnaire_id,xh=xh)
    if len(recorded) != 0:
        if questionnaire_dic.get("oneoff") == 0:
            return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '该问卷一人一份，只能修改单人记录'},
                                json_dumps_params={'ensure_ascii': False})
    return JsonResponse(data={'message': '该学生可以填写问卷'}, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["POST"])
def record_add(request):
    body_list = request_body_serialize_init(request)
    xh = body_list.get('xh')
    questionnaire_id = body_list.get('questionnaireId')
    sub_body_list = list(body_list.values())[list(body_list.keys()).index("v1"):]
    size = len(sub_body_list)
    for i in range(size+1,21):
        sub_body_list.append(None)
    Record.objects.create(xh=xh, questionnaireId=questionnaire_id,
                                 v1=sub_body_list[0], v2=sub_body_list[1], v3=sub_body_list[2], v4=sub_body_list[3],
                                 v5=sub_body_list[4], v6=sub_body_list[5], v7=sub_body_list[6], v8=sub_body_list[7],
                                 v9=sub_body_list[8], v10=sub_body_list[9], v11=sub_body_list[10],
                                 v12=sub_body_list[11],
                                 v13=sub_body_list[12], v14=sub_body_list[13], v15=sub_body_list[14],
                                 v16=sub_body_list[15],
                                 v17=sub_body_list[16], v18=sub_body_list[17], v19=sub_body_list[18],
                                 v20=sub_body_list[19])
    return JsonResponse(data={'message': 'ok'}, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["POST"])
def record_change(request):
    body_list = request_body_serialize_init(request)
    xh = body_list.get('xh')
    questionnaire_id = body_list.get('questionnaireId')
    recs = Record.objects.filter(xh=xh,questionnaireId=questionnaire_id)
    record = Record.objects.get(xh=xh,questionnaireId=questionnaire_id)
    if len(recs) == 0:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该记录（未填写问卷）'},
                            json_dumps_params={'ensure_ascii': False})
    questionnaire_dic = model_to_dict(Questionnaire.objects.get(id=questionnaire_id))
    if questionnaire_dic.get("oneoff") == 1:
        return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '该问卷是一次性问卷，不能修改记录'},
                            json_dumps_params={'ensure_ascii': False})
    for item in body_list.keys():
        if item == "v1":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v1"))
        if item == "v2":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v2"))
        if item == "v3":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v3"))
        if item == "v4":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v4"))
        if item == "v5":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v5"))
        if item == "v6":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v6"))
        if item == "v7":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v7"))
        if item == "v8":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v8"))
        if item == "v9":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v9"))
        if item == "v10":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v10"))
        if item == "v11":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v11"))
        if item == "v12":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v12"))
        if item == "v13":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v13"))
        if item == "v14":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v14"))
        if item == "v15":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v15"))
        if item == "v16":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v16"))
        if item == "v17":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v17"))
        if item == "v18":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v18"))
        if item == "v19":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v19"))
        if item == "v20":
            Record.objects.filter(xh=xh, questionnaireId=questionnaire_id).update(v1=body_list.get("v20"))
    return JsonResponse(data={'message': 'ok'}, json_dumps_params={'ensure_ascii': False})



@permitted_methods(["GET"])
def record_meta(request,recordId):
    record = Record.objects.filter(id=recordId)
    if len(record) == 0 :
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该记录（未填写问卷）'},
                            json_dumps_params={'ensure_ascii': False})
    rec = record[0]
    rec_use = model_to_dict(rec)
    que = model_to_dict(Questionnaire.objects.get(id=rec_use.get("questionnaireId")))
    fields = ["id","xh","questionnaireId"]
    keys = ["id","xh","questionnaireId"]
    for i in range(1, 21):
        tmp = str(i)
        if getattr(rec, "v" + tmp) is not None:
            keys.append(que.get("k" + tmp))
            fields.append("v" + tmp)
    tmp = model_to_dict(rec,fields=fields).values()
    mod = dict(zip(keys,tmp))
    ret = {'message': 'ok','data': mod}
    return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["POST"])
def records(request,questionnaireId):
    page_num = request.GET.get('p', 1)
    length = request.GET.get('l', 5)
    rec = Record.objects.filter(questionnaireId=questionnaireId)
    body_list = request_body_serialize_init(request)
    for item in body_list.keys():
        if item == 'xh' :
            rec = Record.objects.filter(questionnaireId=questionnaireId,xh__icontains=body_list.get('xh'))
        # elif item == 'name':
        #     for ele in rec:
        #         student = Student.objects.filter(xh=ele.xh)
        #         if student.get('name') != body_list.get('name'):
        #             rec.remove(ele)
    paginator = Paginator(rec, length)
    try:
        paginator_page = paginator.page(page_num)
    except EmptyPage:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该页面'}, json_dumps_params={'ensure_ascii': False})
    ret = {'message': 'ok',
           'data': paginator2dict(paginator_page, ["id", "xh", "questionnaireId","v1","v2","v3"])}
    return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["GET"])
def history_meta(request,recordId):
    record = HistoryRecord.objects.filter(id=recordId)
    if len(record) == 0 :
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该归档记录'},
                            json_dumps_params={'ensure_ascii': False})
    rec = record[0]
    rec_use = model_to_dict(rec)
    que = model_to_dict(Questionnaire.objects.get(id=rec_use.get("questionnaireId")))
    fields = ["id","xh","questionnaireId"]
    keys = ["id","xh","questionnaireId"]
    for i in range(1, 21):
        tmp = str(i)
        if getattr(rec, "v" + tmp) is not None:
            keys.append(que.get("k" + tmp))
            fields.append("v" + tmp)
    tmp = model_to_dict(rec,fields=fields).values()
    mod = dict(zip(keys,tmp))
    ret = {'message': 'ok','data': mod}
    return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["GET"])
def history_records(request,questionnaireId):
    page_num = request.GET.get('p', 1)
    length = request.GET.get('l', 5)
    rec = HistoryRecord.objects.filter(questionnaireId=questionnaireId)
    body_list = request_body_serialize_init(request)
    for item in body_list.keys():
        if item == 'xh' :
            rec = HistoryRecord.objects.filter(questionnaireId=questionnaireId,xh__icontains=body_list.get('xh'))
        # elif item == 'name':
        #     for ele in rec:
        #         student = Student.objects.filter(xh=ele.xh)
        #         if student.get('name') != body_list.get('name'):
        #             rec.remove(ele)
    paginator = Paginator(rec, length)
    try:
        paginator_page = paginator.page(page_num)
    except EmptyPage:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该页面'}, json_dumps_params={'ensure_ascii': False})
    ret = {'message': 'ok',
           'data': paginator2dict(paginator_page, ["id", "xh", "questionnaireId","v1","v2","v3"])}
    return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["GET"])
def dynamic_filtering(request,questionnaireId):
    page_num = request.GET.get('p', 1)
    length = request.GET.get('l', 5)
    relations = Whitelist.objects.filter(questionnaireId=questionnaireId)
    xhs = [ele.xh for ele in relations]
    students = Student.objects.filter(xh__in=xhs).order_by('xh')
    # res = model_to_dict(students,fields=['xh','xm','glyx','cc'])
    conditions = Condition.objects.filter(questionnaireId=questionnaireId)
    tmp = []
    contants = " student "
    for ele in conditions:
        tmp.append(model_to_dict(ele,fields=['key','values']))
    for ele in tmp:
        value = "'"+ele['values']+"'"
        l ="SELECT * FROM" + contants + "WHERE"+ele['key']+"="+value
        students = students.raw(l)
        if len(students)==0:
            return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有符合条件的学生'}, json_dumps_params={'ensure_ascii': False})
    #res = []
    #for ele in students:
    #    res.append(model_to_dict(ele,fields=['xh','xm','xq','glyx']))
    paginator = Paginator(students, length)
    try:
        paginator_page = paginator.page(page_num)
    except EmptyPage:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该页面'},
                            json_dumps_params={'ensure_ascii': False})
    ret = {'message': 'ok',
           'data': paginator2dict(paginator_page,
                                  ["xh", "xm", 'glyx', 'xq', ])}
    return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})




@permitted_methods(["POST"])
def file(request,questionnaireId):
    questionnaires = Questionnaire.objects.filter(id=questionnaireId)
    questionnaire = questionnaires[0]
    if questionnaire.status == 1 or questionnaire.status == 2:
        Questionnaire.objects.filter(id=questionnaireId).update(status=-1)
    elif questionnaire.status == 0:
        return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '该问卷仍在草稿状态'},
                            json_dumps_params={'ensure_ascii': False})
    elif questionnaire.status == -1:
        return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '该问卷已归档'},
                            json_dumps_params={'ensure_ascii': False})
    recs = Record.objects.filter(questionnaireId=questionnaireId)
    HistoryRecord.objects.bulk_create(recs)
    Record.objects.filter(questionnaireId=questionnaireId).delete()
    return JsonResponse(data={'message': 'ok'}, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["GET"])
def templateFiles(request, type):
    if type == "whitelist":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment;filename=whitelist.csv'
        response.write(codecs.BOM_UTF8)
        writer = csv.writer(response)
        writer.writerow(["学号"])
        writer.writerow(["10204xxxxx"])
        return response
    elif type == "blacklist":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment;filename=blacklist.csv'
        response.write(codecs.BOM_UTF8)
        writer = csv.writer(response)
        writer.writerow(["学号"])
        writer.writerow(["10204xxxxx"])
        return response
    elif type == "student":
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment;filename=student.csv'
        response.write(codecs.BOM_UTF8)
        writer = csv.writer(response)
        writer.writerow(["学号",'姓名','校区','是否在校','是否在籍','学历层次','管理院系','管理院系码','辅导员姓名','辅导员职工号'])
        writer.writerow(["10204xxxxx",'张三','中北校区','是','是','本科生','软件工程','10010','李四','100300'])
        return response
    else:
        return JsonResponse(status=HTTPStatus.NOT_ACCEPTABLE, data={'error': '参数错误'},
                            json_dumps_params={'ensure_ascii': False})


@permitted_methods(["GET"])
def fileDown(request,questionnaireId):
    questionnaire = Questionnaire.objects.get(id=questionnaireId)
    if questionnaire:
        mes = model_to_dict(questionnaire,fields=['id','status','oneoff','title','scope','creatorId','createTime','updateTime','k1','k2','k3','k4','k5','k6','k7','k8','k9','k10','k11','k12','k13','k14','k15','k16','k17','k18','k19','k20'])
        if mes.get('status') != -1 :
            records = Record.objects.filter(id=questionnaireId)
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment;filename=record.csv'
            response.write(codecs.BOM_UTF8)
            writer = csv.writer(response)
            writer.writerow(["序号","学号", "创建时间", "更新时间", mes.get('k1'), mes.get('k2'), mes.get('k3'), mes.get('k4'), mes.get('k4'), mes.get('k5'), mes.get('k6'), mes.get('k7'), mes.get('k8'), mes.get('k9'), mes.get('k10'),
                   mes.get('k11'),mes.get('k12'),mes.get('k13'),mes.get('k14'),mes.get('k15'),mes.get('k16'),mes.get('k17'),mes.get('k18'),mes.get('k19'),mes.get('k20')])
            j = 1
            for record in records:
                tmp = model_to_dict(record,fields=['xh','createTime','updateTime','v1','v2','v3','v4','v5','v6','v7','v8','v9','v10','v11','v12','v13','v14','v15','v16','v17','v18','v19','v20'])
                tmp['createTime'] = record.createTime.strftime("%Y-%m-%d %H:%M")
                tmp['updateTime'] = record.updateTime.strftime("%Y-%m-%d %H:%M")
                writer.writerow([j,tmp.get('xh'),tmp.get('createTime'),tmp.get('updateTime'),tmp.get('v1'),tmp.get('v2'),tmp.get('v3'),tmp.get('v4'),tmp.get('v5'),tmp.get('v6'),tmp.get('v7'),tmp.get('v8'),tmp.get('v8'),tmp.get('v9'),tmp.get('v9'),tmp.get('v10'),tmp.get('v11'),tmp.get('v12'),tmp.get('v13'),tmp.get('v14'),tmp.get('v15'),tmp.get('v16'),tmp.get('v17'),tmp.get('v18'),tmp.get('v19'),tmp.get('v20')])
                j = j+1
        else:
            history_records = HistoryRecord.objects.filter(questionnaireId=questionnaireId)
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment;filename=history_record.csv'
            response.write(codecs.BOM_UTF8)
            writer = csv.writer(response)
            writer.writerow(
                ["序号", "学号", "创建时间", "更新时间",mes.get('k1'), mes.get('k2'), mes.get('k3'), mes.get('k4'), mes.get('k4'),
                 mes.get('k5'), mes.get('k6'), mes.get('k7'), mes.get('k8'), mes.get('k9'), mes.get('k10'),
                 mes.get('k11'), mes.get('k12'), mes.get('k13'), mes.get('k14'), mes.get('k15'), mes.get('k16'),
                 mes.get('k17'), mes.get('k18'), mes.get('k19'), mes.get('k20')])
            j = 1
            for history_record in history_records:
                tmp = model_to_dict(history_record,
                                    fields=['xh', 'createTime', 'updateTime', 'v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7',
                                            'v8', 'v9', 'v10', 'v11', 'v12', 'v13', 'v14', 'v15', 'v16', 'v17', 'v18',
                                            'v19', 'v20'])
                tmp['createTime'] = history_record.createTime.strftime("%Y-%m-%d %H:%M")
                tmp['updateTime'] = history_record.updateTime.strftime("%Y-%m-%d %H:%M")
                writer.writerow(
                    [j, tmp.get('xh'), tmp.get('createTime'), tmp.get('updateTime'), tmp.get('v1'), tmp.get('v2'),
                     tmp.get('v3'), tmp.get('v4'), tmp.get('v5'), tmp.get('v6'), tmp.get('v7'), tmp.get('v8'),
                     tmp.get('v8'), tmp.get('v9'), tmp.get('v9'), tmp.get('v10'), tmp.get('v11'), tmp.get('v12'),
                     tmp.get('v13'), tmp.get('v14'), tmp.get('v15'), tmp.get('v16'), tmp.get('v17'), tmp.get('v18'),
                     tmp.get('v19'), tmp.get('v20')])
                j = j + 1
    return response

class AnswerUserLogin(wx.Frame):

    def __init__(self, *args, **kw):
        super(AnswerUserLogin, self).__init__(*args, **kw)
        self.Center()
        self.pnl = wx.Panel(self)
        self.LoginInterface()

    def LoginInterface(self):
        self.vbox = wx.BoxSizer(wx.VERTICAL)
        logo = wx.StaticText(self.pnl, label="问卷系统")
        font = logo.GetFont()
        font.PointSize += 30
        font = font.Bold()
        logo.SetFont(font)
        self.vbox.Add(logo, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=150)

        Id = wx.StaticBox(self.pnl, label="学号")
        hobox = wx.StaticBoxSizer(Id, wx.HORIZONTAL)
        self.userId = wx.TextCtrl(self.pnl, size=(300, 25))
        hobox.Add(self.userId, 0, wx.EXPAND | wx.BOTTOM, 5)
        self.vbox.Add(hobox,  proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=20)

        login_button = wx.Button(self.pnl, label="登录", size=(80, 25))
        login_button.Bind(wx.EVT_BUTTON, self.LoginButton)
        self.vbox.Add(login_button, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=20)

        self.pnl.SetSizer(self.vbox)

    def LoginButton(self, event):
        userId=self.userId.GetValue()
        ret=Student.objects.filter(xh=userId)
        if len(ret)==0:
            warning=wx.StaticText(self.pnl, label="学号错误")
            warning.SetForegroundColour(wx.RED)
            self.vbox.Add(warning, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=20)
        else:
            globalVariable.user_id=userId
            if globalVariable.typesList[0] == 0:
                operate = AnswerChoice(None, title="填写问卷", size=(1024, 668))
                operate.Show()
            if globalVariable.typesList[0] == 1:
                operate = AnswerFill(None, title="填写问卷", size=(1024, 668))
                operate.Show()
            self.Close(True)

def string_after_deal(topic):
    length = len(topic)
    if length > 25:
        ret = topic[0:25] + '\n' + topic[25:length]
    return ret

class AnswerFill(wx.Frame):
    def __init__(self,*args, **kw):
        super(AnswerFill, self).__init__(*args, **kw)
        self.Center()
        self.pnl = wx.Panel(self)
        self.topic = globalVariable.get_current()

        self.vbox = wx.BoxSizer(wx.VERTICAL)

        logo = wx.StaticText(self.pnl, label="问卷填写")
        font = logo.GetFont()
        font.PointSize += 30
        font = font.Bold()
        logo.SetFont(font)
        self.vbox.Add(logo, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=50)

        notice = wx.StaticBox(self.pnl,
                              label="题目" + str(globalVariable.count + 1) + '/' + str(len(globalVariable.typesList)))
        self.vbox.Add(notice, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=30)


        que = wx.StaticBox(self.pnl, label="问 题")
        hbox = wx.StaticBoxSizer(que, wx.HORIZONTAL)
        self.question = wx.StaticText(self.pnl, -1, string_after_deal(self.topic[0]), size=(700, 100), style=wx.ALIGN_CENTER)
        font = wx.Font(wx.FontInfo(15).Bold())
        self.question.SetFont(font)
        hbox.Add(self.question, 0, wx.EXPAND | wx.TOP, 5)
        self.vbox.Add(hbox, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=50)


        warning = wx.StaticText(self.pnl, -1, '回答最多填写50个字', size=(200, 50), style=wx.ALIGN_CENTER)
        self.vbox.Add(warning, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=10)

        self.multiText = wx.TextCtrl(self.pnl, -1, "", size=(700, 100), style=wx.TE_MULTILINE)
        self.multiText.SetInsertionPoint(0)
        self.multiText.SetMaxLength(50)
        font = wx.Font(wx.FontInfo(15))
        self.multiText.SetFont(font)
        self.vbox.Add(self.multiText, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=10)

        self.sure = wx.Button(self.pnl, label="提交", size=(80, 25))
        self.sure.Bind(wx.EVT_BUTTON,self.submit)
        self.vbox.Add(self.sure, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=10)

        self.pnl.SetSizer(self.vbox)

    def submit(self, event):
        ans = self.multiText.GetValue()
        globalVariable.append_answer(ans)
        nextType =globalVariable.next_type()
        if nextType==-1:
            self.Close(True)
        if nextType==0:
            operation = AnswerChoice(None, title="问卷填写", size=(1024, 668))
            operation.Show()
            self.Close(True)
        if nextType==1:
            operation = AnswerFill(None,title="问卷填写", size=(1024, 668))
            operation.Show()
            self.Close(True)







class AnswerChoice(wx.Frame):


    def __init__(self,*args, **kw):
        super(AnswerChoice, self).__init__(*args, **kw)
        self.Center()
        self.pnl = wx.Panel(self)
        self.vbox = wx.BoxSizer(wx.VERTICAL)
        self.topic = globalVariable.get_current()

        logo = wx.StaticText(self.pnl, label="问卷填写")
        font = logo.GetFont()
        font.PointSize += 30
        font = font.Bold()
        logo.SetFont(font)
        self.vbox.Add(logo, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=5)

        notice = wx.StaticBox(self.pnl, label="题目"+str(globalVariable.count+1)+'/'+str(len(globalVariable.typesList)))
        self.vbox.Add(notice, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=30)

        self.question = wx.StaticText(self.pnl, -1,string_after_deal(self.topic[0]),size = (700, 50), style=wx.ALIGN_CENTER)  # 创建一个文本控件
        self.choiceA = wx.StaticText(self.pnl, -1, string_after_deal(self.topic[1]), size=(700, 50), style=wx.ALIGN_CENTER)  # 创建一个文本控件
        self.choiceB = wx.StaticText(self.pnl, -1, string_after_deal(self.topic[2]), size=(700, 50), style=wx.ALIGN_CENTER)  # 创建一个文本控件
        self.choiceC = wx.StaticText(self.pnl, -1, string_after_deal(self.topic[3]), size=(700, 50), style=wx.ALIGN_CENTER)  # 创建一个文本控件
        self.choiceD = wx.StaticText(self.pnl, -1, string_after_deal(self.topic[4]), size=(700, 50), style=wx.ALIGN_CENTER)  # 创建一个文本控件
        font = wx.Font(wx.FontInfo(15).Bold() )
        self.question.SetFont(font)
        self.choiceA.SetFont(font)
        self.choiceB.SetFont(font)
        self.choiceC.SetFont(font)
        self.choiceD.SetFont(font)


        que = wx.StaticBox(self.pnl, label="问 题")
        cha = wx.StaticBox(self.pnl, label="选项A")
        chb = wx.StaticBox(self.pnl, label="选项B")
        chc = wx.StaticBox(self.pnl, label="选项C")
        chd = wx.StaticBox(self.pnl, label="选项D")

        hobox_que = wx.StaticBoxSizer(que, wx.HORIZONTAL)
        hobox_cha = wx.StaticBoxSizer(cha, wx.HORIZONTAL)
        hobox_chb = wx.StaticBoxSizer(chb, wx.HORIZONTAL)
        hobox_chc = wx.StaticBoxSizer(chc, wx.HORIZONTAL)
        hobox_chd = wx.StaticBoxSizer(chd, wx.HORIZONTAL)

        hobox_que.Add(self.question, 0, wx.EXPAND | wx.TOP, 5)
        hobox_cha.Add(self.choiceA, 0, wx.EXPAND | wx.TOP, 5)
        hobox_chb.Add(self.choiceB, 0, wx.EXPAND | wx.TOP, 5)
        hobox_chc.Add(self.choiceC, 0, wx.EXPAND | wx.TOP, 5)
        hobox_chd.Add(self.choiceD, 0, wx.EXPAND | wx.TOP, 5)

        self.vbox.Add(hobox_que, 0, wx.CENTER | wx.FIXED_MINSIZE, 5)
        self.vbox.Add(hobox_cha, 0, wx.CENTER | wx.FIXED_MINSIZE, 5)
        self.vbox.Add(hobox_chb, 0, wx.CENTER | wx.FIXED_MINSIZE, 5)
        self.vbox.Add(hobox_chc, 0, wx.CENTER | wx.FIXED_MINSIZE, 5)
        self.vbox.Add(hobox_chd, 0, wx.CENTER | wx.FIXED_MINSIZE, 5)


        choiceList = ['A', 'B', 'C', 'D']
        self.radioBox = wx.RadioBox(self.pnl, -1, "", wx.DefaultPosition, wx.DefaultSize,choiceList, 4, wx.RA_SPECIFY_COLS)
        self.radioBox.SetSelection(0)
        self.vbox.Add(self.radioBox,  proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=10)

        self.sure = wx.Button(self.pnl, label="提交", size=(80, 25))
        self.sure.Bind(wx.EVT_BUTTON, self.Submit)
        self.vbox.Add(self.sure, 0, wx.CENTER | wx.FIXED_MINSIZE, 5)

        self.pnl.SetSizer(self.vbox)

    def Submit(self, event):
        selected = self.radioBox.GetSelection()
        if selected == 0:
            ans=self.choiceA.GetLabel()
        if selected==1:
            ans=self.choiceB.GetLabel()
        if selected==2:
            ans=self.choiceC.GetLabel()
        if selected==3:
            ans=self.choiceD.GetLabel()
        globalVariable.append_answer(ans)
        nextType = globalVariable.next_type()
        if nextType == -1:
            self.Close(True)
        if nextType == 0:
            operation = AnswerChoice(None ,title="问卷填写", size=(1024, 668))
            operation.Show()
            self.Close(True)
        if nextType == 1:
            operation = AnswerFill(None,title="问卷填写", size=(1024, 668))
            operation.Show()
            self.Close(True)


class globalVariable:

    answers=[]
    typesList=[]
    titlesList=[]
    count=0
    user_id=None

    def __init__(self):
        pass

    @classmethod
    def next_type(cls):
        cls.count += 1
        if cls.count >= len(cls.typesList):
            return -1
        else:
            return cls.typesList[cls.count]

    @classmethod
    def get_current(cls):
        print( cls.titlesList[cls.count])
        return cls.titlesList[cls.count]

    @classmethod
    def append_answer(cls,ans):
        cls.answers.append(ans)

    @classmethod
    def answer_padding(cls):
        while len(cls.answers)<20:
            cls.answers.append(None)

    @classmethod
    def get_answer(cls):
        return cls.answers

def End():
    ret = wx.MessageBox('您已完成问卷\n点击确认提交', 'warning',wx.OK | wx.CANCEL)
    return ret

#回答问卷接口
@permitted_methods(["GET"])
def answer(request, pollId):
    questionnaires = Questionnaire.objects.filter(id=pollId)
    if len(questionnaires) == 0:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该问卷'},
                            json_dumps_params={'ensure_ascii': False})
    choice = Choice.objects.filter(questionnaireId=pollId).order_by('choiceNumber')
    choices = []
    for item in choice:
        choices.append(model_to_dict(item))

    questionnaire = questionnaires[0]
    questionnaire_dict = model_to_dict(questionnaire)
    types = []
    titles = []
    for i in range(20):
        field = 'type' + str(i + 1)
        if questionnaire_dict[field] == 2 or questionnaire_dict[field] == None:
            break
        else:
            types.append(questionnaire_dict[field])
            if questionnaire_dict[field] == 1:
                field = 'k' + str(i + 1)
                title=[questionnaire_dict[field]]
            elif questionnaire_dict[field] == 0:
                field = 'k' + str(i + 1)
                title = [questionnaire_dict[field]]
                for item in choices:
                    if item['choiceNumber'] == i + 1:
                        title.append(item['A'])
                        title.append(item['B'])
                        title.append(item['C'])
                        title.append(item['D'])
                        break
            titles.append(title)
    globalVariable.count=0
    globalVariable.answers=[]
    globalVariable.typesList=types
    globalVariable.titlesList=titles

    if len(types) > 0:
        app = wx.App()
        operate=AnswerUserLogin(None,title="填写问卷", size=(1100, 700))
        operate.Show()
        app.MainLoop()

    if len(globalVariable.answers)<len(globalVariable.typesList):
        ret = {'message': 'cancel'}
        return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})
    retValue = End()
    if retValue == wx.OK:
        globalVariable.answer_padding()
        answers = globalVariable.get_answer()
        Record.objects.create(xh=globalVariable.user_id, questionnaireId=pollId, v1=answers[0], v2=answers[1], v3=answers[2], v4=answers[3],
                              v5=answers[4], v6=answers[5], v7=answers[6], v8=answers[7], v9=answers[8], v10=answers[9],
                              v11=answers[10], v12=answers[11], v13=answers[12], v14=answers[13], v15=answers[14],
                              v16=answers[15], v17=answers[16], v18=answers[17], v19=answers[18], v20=answers[19])
        ret = {'message': 'complete'}
        return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})
    else:
        ret = {'message': 'cancel'}
        return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})


class CreateUserLogin(wx.Frame):

    def __init__(self, *args, **kw):
        super(CreateUserLogin, self).__init__(*args, **kw)
        self.Center()
        self.pnl = wx.Panel(self)
        self.vbox = wx.BoxSizer(wx.VERTICAL)
        self.LoginInterface()


    def LoginInterface(self):

        logo = wx.StaticText(self.pnl, label="问卷系统")
        font = logo.GetFont()
        font.PointSize += 30
        font = font.Bold()
        logo.SetFont(font)
        self.vbox.Add(logo, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=150)

        Id = wx.StaticBox(self.pnl, label="学号")
        hobox = wx.StaticBoxSizer(Id, wx.HORIZONTAL)
        self.userId = wx.TextCtrl(self.pnl, size=(300, 25))
        hobox.Add(self.userId, 0, wx.EXPAND | wx.BOTTOM, 5)
        self.vbox.Add(hobox,  proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=20)

        login_button = wx.Button(self.pnl, label="登录", size=(80, 25))
        login_button.Bind(wx.EVT_BUTTON, self.LoginButton)
        self.vbox.Add(login_button, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=20)

        self.pnl.SetSizer(self.vbox)

    def LoginButton(self, event):
        userId = self.userId.GetValue()
        ret = Student.objects.filter(xh=userId)
        if len(ret) == 0:
            warning = wx.StaticText(self.pnl, label="学号错误")
            warning.SetForegroundColour(wx.RED)
            self.vbox.Add(warning, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=20)
        else:
            Total.user_id=userId
            operation = UserOperation(None, title="问卷系统", size=(1100, 700))
            operation.Show()
            self.Close(True)


class UserOperation(wx.Frame):

    def __init__(self, *args, **kw):
        super(UserOperation, self).__init__(*args, **kw)
        self.Center()
        self.pnl = wx.Panel(self)
        self.vbox = wx.BoxSizer(wx.VERTICAL)

        logo = wx.StaticText(self.pnl, label="设计问卷")
        self.vbox.Add(logo, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=5)

        buttonBox = wx.StaticBox(self.pnl, label="")
        buttonRow = wx.StaticBoxSizer(buttonBox, wx.HORIZONTAL)

        choice_button = wx.Button(self.pnl, id=10, label="选择题", size=(100, 50))
        fill_button = wx.Button(self.pnl, id=11, label="填空题", size=(100, 50))
        check_button = wx.Button(self.pnl, id=12, label="查看问卷", size=(100, 50))
        submit_button = wx.Button(self.pnl, id=13, label="提交问卷", size=(100, 50))

        buttonRow.Add(choice_button, 0, wx.EXPAND | wx.BOTTOM, 10)
        buttonRow.Add(fill_button, 0, wx.EXPAND | wx.BOTTOM, 10)
        buttonRow.Add(check_button, 0, wx.EXPAND | wx.BOTTOM, 10)
        buttonRow.Add(submit_button, 0, wx.EXPAND | wx.BOTTOM, 10)
        self.vbox.Add(buttonRow, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=0)

        self.Bind(wx.EVT_BUTTON, self.ClickButton, id=10, id2=13)

        window = wx.StaticBox(self.pnl, label="",size=(800,500))
        self.window_box_sizer = wx.StaticBoxSizer(window ,wx.VERTICAL)
        self.vbox.Add(self.window_box_sizer, proportion=0, flag=wx.CENTER)

        self.pnl.SetSizer(self.vbox)

    def ClickButton(self, event):
        source_id = event.GetId()
        if source_id == 10:
            print("choice_button！")
            choice_interfece = ChoiceOp(None, title="设计选择题", size=(1024, 668))
            choice_interfece.Show()
            self.Close(True)
        elif source_id == 11:
            print("fill_button！")
            fill_button = FillOp(None, title="设计填空题", size=(1024, 668))
            fill_button.Show()
            self.Close(True)
        elif source_id == 12:
            print("check_button！")
            check_button = CheckOp(None, title="查看问卷题目", size=(1024, 668))
            check_button.Show()
            self.Close(True)
        elif source_id==13:
            if submit_to_database()==wx.OK:
                self.Destroy()

class Total():
    total_title_number=0
    total_tiltes=[]
    user_id=None
    def __init__(self):
        pass

    @classmethod
    def add_title(cls,title):
        cls.total_title_number+=1
        cls.total_tiltes.append(title)

    @classmethod
    def get_content(cls,row,col):
        if col==0:
            if len(cls.total_tiltes[row])==1:
                return'填空'
            elif len(cls.total_tiltes[row])!=1:
                return'选择'
        elif col<=len(cls.total_tiltes[row]):
            return cls.total_tiltes[row][col-1]
        else:
            return''




class ChoiceOp(UserOperation):
    def __init__(self, *args, **kw):
        super(ChoiceOp, self).__init__(*args, **kw)

        warning = wx.StaticText(self.pnl, -1, '题目和选项在50字以内', size=(100, 20), style=wx.ALIGN_CENTER)
        self.window_box_sizer.Add(warning, 0, wx.CENTER | wx.TOP | wx.FIXED_MINSIZE, 5)

        self.title = wx.TextCtrl(self.pnl, size=(700, 50))
        self.A = wx.TextCtrl(self.pnl, size=(700, 50))
        self.B = wx.TextCtrl(self.pnl, size=(700, 50))
        self.C = wx.TextCtrl(self.pnl, size=(700, 50))
        self.D = wx.TextCtrl(self.pnl, size=(700, 50))
        self.title.SetMaxLength(50)
        self.A.SetMaxLength(50)
        self.B.SetMaxLength(50)
        self.C.SetMaxLength(50)
        self.D.SetMaxLength(50)
        self.add = wx.Button(self.pnl,label="确认添加",size=(80,25))
        self.add.Bind(wx.EVT_BUTTON, self.Add)

        title_box = wx.StaticBox(self.pnl, label="题 目")
        Abox = wx.StaticBox(self.pnl, label="选择A")
        Bbox = wx.StaticBox(self.pnl, label="选择B")
        Cbox = wx.StaticBox(self.pnl, label="选择C")
        Dbox = wx.StaticBox(self.pnl, label="选择D")

        hbox_title = wx.StaticBoxSizer(title_box, wx.HORIZONTAL)
        hbox_A = wx.StaticBoxSizer(Abox, wx.HORIZONTAL)
        hbox_B = wx.StaticBoxSizer(Bbox, wx.HORIZONTAL)
        hbox_C = wx.StaticBoxSizer(Cbox, wx.HORIZONTAL)
        hbox_D = wx.StaticBoxSizer(Dbox, wx.HORIZONTAL)

        hbox_title.Add(self.title, 0, wx.EXPAND | wx.BOTTOM, 5)
        hbox_A.Add(self.A, 0, wx.EXPAND | wx.BOTTOM, 5)
        hbox_B.Add(self.B, 0, wx.EXPAND | wx.BOTTOM, 5)
        hbox_C.Add(self.C, 0, wx.EXPAND | wx.BOTTOM, 5)
        hbox_D.Add(self.D, 0, wx.EXPAND | wx.BOTTOM, 5)

        self.window_box_sizer.Add(hbox_title, 0, wx.CENTER | wx.TOP | wx.FIXED_MINSIZE, 5)
        self.window_box_sizer.Add(hbox_A, 0, wx.CENTER | wx.TOP | wx.FIXED_MINSIZE, 5)
        self.window_box_sizer.Add(hbox_B, 0, wx.CENTER | wx.TOP | wx.FIXED_MINSIZE, 5)
        self.window_box_sizer.Add(hbox_C, 0, wx.CENTER | wx.TOP | wx.FIXED_MINSIZE, 5)
        self.window_box_sizer.Add(hbox_D, 0, wx.CENTER | wx.TOP | wx.FIXED_MINSIZE, 5)
        self.window_box_sizer.Add(self.add, 0, wx.CENTER | wx.TOP | wx.FIXED_MINSIZE, 5)

    def ClickButton(self, event):
        source_id = event.GetId()
        if source_id == 10:
            pass
        elif source_id == 11:
            print("fill_button！")
            fill_button = FillOp(None, title="设计填空题", size=(1024, 668))
            fill_button.Show()
            self.Close(True)
        elif source_id == 12:
            print("check_button！")
            check_button = CheckOp(None, title="查看问卷题目", size=(1024, 668))
            check_button.Show()
            self.Close(True)
        elif source_id==13:
            if submit_to_database() == wx.OK:
                self.Destroy()


    def Add(self, event):
        Total.add_title([self.title.GetValue(),self.A.GetValue(),self.B.GetValue(),self.C.GetValue(),self.D.GetValue()])
        print([self.title.GetValue(),self.A.GetValue(),self.B.GetValue(),self.C.GetValue(),self.D.GetValue()])
        self.title.SetValue('')
        self.A.SetValue('')
        self.B.SetValue('')
        self.C.SetValue('')
        self.D.SetValue('')


class FillOp(UserOperation):
    def __init__(self, *args, **kw):
        super(FillOp, self).__init__(*args, **kw)
        self.title = wx.TextCtrl(self.pnl, pos=(407, 78), size=(700, 200))
        self.title.SetMaxLength(50)

        self.add = wx.Button(self.pnl, label="确认添加", pos=(625, 78), size=(80, 25))
        self.add.Bind(wx.EVT_BUTTON, self.Add)

        title_box = wx.StaticBox(self.pnl, label="请输入题目")

        hobox = wx.StaticBoxSizer(title_box, wx.HORIZONTAL)

        hobox.Add(self.title, 0, wx.EXPAND | wx.BOTTOM, 5)

        warning = wx.StaticText(self.pnl, -1, '题目在50字以内', size=(100, 20), style=wx.ALIGN_CENTER)
        self.window_box_sizer.Add(warning, 0, wx.CENTER | wx.TOP | wx.FIXED_MINSIZE, 5)
        self.window_box_sizer.Add(hobox, 0, wx.CENTER | wx.TOP | wx.FIXED_MINSIZE, 5)
        self.window_box_sizer.Add(self.add, 0, wx.CENTER | wx.TOP | wx.FIXED_MINSIZE, 5)

    def ClickButton(self, event):
        source_id = event.GetId()
        if source_id == 10:
            print("choice_button！")
            choice_interfece = ChoiceOp(None, title="设计选择题", size=(1024, 668))
            choice_interfece.Show()
            self.Close(True)
        elif source_id == 11:
            pass
        elif source_id == 12:
            print("check_button！")
            check_button = CheckOp(None, title="查看问卷题目", size=(1024, 668))
            check_button.Show()
            self.Close(True)
        elif source_id == 13:
            if submit_to_database() == wx.OK:
                self.Destroy()


    def Add(self, event):
        Total.add_title([self.title.GetValue()])
        print([self.title.GetValue()])
        self.title.SetValue('')



class CheckOp(UserOperation):
    def __init__(self, *args, **kw):
        super(CheckOp, self).__init__(*args, **kw)
        self.titles_grid = self.CreateTitlesGrid()
        self.window_box_sizer.Add(self.titles_grid, 0, wx.CENTER | wx.TOP, 10)
        self.Bind(wx.grid.EVT_GRID_SELECT_CELL, self.select,self.titles_grid)

        self.text=wx.TextCtrl(self.pnl, size=(700, 50))
        self.vbox.Add(self.text, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=5)

        # button_row=wx.StaticBox(self.pnl, label="操作")
        # self.delete_button=wx.Button(self.pnl, label="删除此题",size=(80, 25))
        # self.update_button=wx.Button(self.pnl, label="修改此题",size=(80, 25))
        # self.delete_button.Bind(wx.EVT_BUTTON, self.delete)
        # self.update_button.Bind(wx.EVT_BUTTON, self.update)
        # hbox = wx.StaticBoxSizer(button_row, wx.HORIZONTAL)
        # hbox.Add(self.delete_button,0,wx.EXPAND | wx.BOTTOM|wx.CENTER, 5)
        # hbox.Add(self.update_button, 0, wx.EXPAND | wx.BOTTOM|wx.CENTER, 5)
        # self.vbox.Add(hbox, proportion=0, flag=wx.FIXED_MINSIZE | wx.TOP | wx.CENTER, border=0)


    def ClickButton(self, event):
        source_id = event.GetId()
        if source_id == 10:
            print("choice_button！")
            choice_interfece = ChoiceOp(None, title="设计选择题", size=(1024, 668))
            choice_interfece.Show()
            self.Close(True)
        elif source_id == 11:
            print("fill_button！")
            fill_button = FillOp(None, title="设计填空题", size=(1024, 668))
            fill_button.Show()
            self.Close(True)
        elif source_id == 12:
            pass
        elif source_id==13:
            if(submit_to_database()==wx.OK):
                self.Destroy()

    def CreateTitlesGrid(self):
        titles_gird = wx.grid.Grid(self.pnl,0,  size=(850,350),style=wx.WANTS_CHARS, name='')
        lines=Total.total_title_number
        titles_gird.CreateGrid(lines, 6)
        column_names = ("题型","题目", "A", "B", "C", "D")
        for i in range(6):
            titles_gird.SetColLabelValue(i,column_names[i])
        titles_gird.SetCornerLabelValue('题号')
        titles_gird.SetDefaultColSize(140, resizeExistingCols=False)
        titles_gird.SetColSize(0,50)
        titles_gird.SetDefaultRowSize(20, resizeExistingRows=False)
        for i in range(lines):
            for j in range(6):
                titles_gird.SetCellValue(row=i,col=j,s=Total.get_content(i,j))
        return titles_gird

    def select(self, event):
        c=event.GetCol()
        r=event.GetRow()
        value=self.titles_grid.GetCellValue(row=r,col=c)
        print(value)
        self.text.SetValue(Total.get_content(r,c))

def submit_to_database():
    if(Total.total_title_number==0):
        ret = wx.MessageBox('尚未创建题目\n是否确认提交', 'warning',wx.OK | wx.CANCEL)
        if ret==wx.OK:
            Questionnaire.objects.create(status=-1, oneoff=1, title='no', scope=1, creatorId=Total.user_id)
    else:
        ret = wx.MessageBox('是否确认提交', 'warning',wx.OK | wx.CANCEL)
        if ret==wx.OK:
            types = []
            for i in range(Total.total_title_number):
                if len(Total.total_tiltes[i]) == 1:
                    types.append(1)
                else:
                    types.append(0)
            for i in range(Total.total_title_number,20):
                Total.add_title([None])
                types.append(2)
            Questionnaire.objects.create(status=-1, oneoff=1, title='no', scope=1, creatorId=Total.user_id,
                                        k1=Total.total_tiltes[0][0],k2=Total.total_tiltes[1][0],k3=Total.total_tiltes[2][0],
                                        k4=Total.total_tiltes[3][0],k5=Total.total_tiltes[4][0],k6=Total.total_tiltes[5][0],
                                        k7=Total.total_tiltes[6][0],k8=Total.total_tiltes[7][0],k9=Total.total_tiltes[8][0],
                                        k10=Total.total_tiltes[9][0],k11=Total.total_tiltes[10][0],k12=Total.total_tiltes[11][0],
                                        k13=Total.total_tiltes[12][0],k14=Total.total_tiltes[13][0],k15=Total.total_tiltes[14][0],
                                        k16=Total.total_tiltes[15][0],k17=Total.total_tiltes[16][0],k18=Total.total_tiltes[17][0],
                                        k19=Total.total_tiltes[18][0],k20=Total.total_tiltes[19][0],type1=types[0],type2=types[1],
                                        type3=types[2],type4=types[3],type5=types[4],type6=types[5],type7=types[6],type8=types[7],
                                        type9=types[8],type10=types[9],type11=types[10],type12=types[11],type13=types[12],
                                        type14=types[13],type15=types[14],type16=types[15],type17=types[16],type18=types[17],
                                         type19=types[18],type20=types[19],
)
            questionnaires=Questionnaire.objects.filter(creatorId=Total.user_id).order_by('createTime')
            questionnaire =model_to_dict(questionnaires[len(questionnaires)-1])
            for i in range(Total.total_title_number):
                if len(Total.total_tiltes[i])!=1:
                    Choice.objects.create(questionnaireId=questionnaire['id'],choiceNumber=i+1,A=Total.total_tiltes[i][1],
                                          B=Total.total_tiltes[i][2],C=Total.total_tiltes[i][3],D=Total.total_tiltes[i][4])
    return ret


#创建问卷接口
@permitted_methods(["GET"])
def create(request):
    app = wx.App()
    operate = CreateUserLogin(None, title="填写问卷", size=(1100, 700))
    operate.Show()
    app.MainLoop()
    ret = {'message': 'ok'}
    return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})







