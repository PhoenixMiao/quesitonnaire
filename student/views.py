import datetime

from django.shortcuts import render

from Utils.wrappers import permitted_methods
from student.models import Student
from student.models import Instructor_Student
from django.http import JsonResponse
from django.core.paginator import Paginator, EmptyPage
from django.core import serializers
from Utils.tools import paginator2dict,paginator3dict
from Utils.midd import errMsg
from http import HTTPStatus
from Utils.tools import request_body_serialize_init
from .models import DepartAdmin
from django.forms.models import model_to_dict
from Utils.fileTool import upload_file2
from openpyxl import Workbook
from openpyxl import load_workbook

@permitted_methods(["POST"])
def student(request):
    stu_mes = request_body_serialize_init(request)
    userId = '20130053'
    page_num = request.GET.get('p', 1)
    length = request.GET.get('l', 5)
#    permission = '4'
    relations = Instructor_Student.objects.filter(zgh=userId)
    xhs = [ele.xh for ele in relations]
    students = Student.objects.filter(xh__in=xhs).order_by('xh')
    for item in stu_mes.keys():
        if item == "xm":
            students  = students.filter(xm__icontains=stu_mes.get(item))
        elif item == "xh":
            students = students.filter(xh__icontains=stu_mes.get(item))
        elif item == "xq":
            students  = students.filter(xq__icontains=stu_mes.get(item))
        elif item == "sfzx":
            students  = students.filter(sfzx__icontains=stu_mes.get(item))
        elif item == "cc":
            students  = students.filter(cc__icontains=stu_mes.get(item))
        elif item == "glyx":
            students  = students.filter(glyx__icontains=stu_mes.get(item))
    paginator = Paginator(students, length)
    try:
        paginator_page = paginator.page(page_num)
    except EmptyPage:
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '没有该页面'}, json_dumps_params={'ensure_ascii': False})
    ret = {'message': 'ok',
           'data': paginator3dict(paginator_page, ["xh", "xm", 'glyx', 'glyxm', 'sftb', 'sfdr', 'sfzj', 'sfzx', 'xq', ])}
    return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})

@permitted_methods(["POST"])
def meta(request):
    stu_mes = request_body_serialize_init(request)
    student = Student.objects.get(xh=stu_mes['xh'])
    mod = model_to_dict(student)
    mod2 = model_to_dict(Instructor_Student.objects.get(xh=mod['xh']))
    mod3 = model_to_dict(DepartAdmin.objects.get(glyx=mod.get('glyx')))
    mod['instructor_zgh'] = mod2.get('zgh')
    mod['instructor_xm'] = mod2.get('xm')
    mod['depart_admin_zgh'] = mod3.get('zgh')
    mod['depart_admin_xm'] = mod3.get('xm')
    if mod['sftb'] == True:
        mod['sftb'] = '是'
    if mod['sftb'] == False:
        mod['sftb'] = '否'
    if mod['sfdr'] == True:
        mod['sfdr'] = '是'
    if mod['sfdr'] == False:
        mod['sfdr'] = '否'
    ret = {'message': 'ok',
           'data': mod}
    return JsonResponse(data=ret, json_dumps_params={'ensure_ascii': False})

@permitted_methods(["POST"])
def student_import(request):
    tmpFile = request.FILES.get('studentlist')
    workbook_ = load_workbook(tmpFile)
    sheetnames = workbook_.get_sheet_names()
    sheet = workbook_.get_sheet_by_name(sheetnames[0])
    j=2
    tmp = sheet.cell(row=2,column=1).value
    while tmp != None:
        stu_mes = {}
        title = ['xh','name','xq','sfzx','sfzj','cc','glyx','glyxm','instructor_name','instructor_num']
        for i in range(0,10):
            stu_mes[title[i]] = sheet.cell(row=j,column=i+1).value
        relations2 = Student.objects.filter(xh=stu_mes.get("xh"))
        if len(relations2) >0:
            #return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '该学生已存在'},json_dumps_params={'ensure_ascii': False})
            continue
        if stu_mes.get("xh")==None or stu_mes.get("name")==None or stu_mes.get("xq")==None or stu_mes.get("sfzx")==None or stu_mes.get("cc")==None or stu_mes.get("glyx")==None or stu_mes.get("instructor_name")==None or stu_mes.get("instructor_num")==None:
            #return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '参数不全，字段不齐'},json_dumps_params={'ensure_ascii': False})
            continue
        Student.objects.create(xh=stu_mes.get("xh"),xm=stu_mes.get("name"),xq=stu_mes.get("xq"),sfzx=stu_mes.get("sfzx"),sfzj=stu_mes.get('sfzj'),
                            cc=stu_mes.get("cc"),glyx=stu_mes.get("glyx"),glyxm=stu_mes.get('glyxm'),sfdr=True,sftb=False)
        Instructor_Student.objects.create(zgh=stu_mes.get("instructor_num"),xm=stu_mes.get("instructor_name"),xh=stu_mes.get("xh"))
        j = j+1
        tmp = sheet.cell(row=j,column=1).value
    return JsonResponse(data={'message': 'ok'}, json_dumps_params={'ensure_ascii': False})


@permitted_methods(["POST"])
def student_change(request):
    stus = upload_file
    stu_mes = request_body_serialize_init(request)
    relations = Student.objects.filter(xh=stu_mes.get("xh"))
    if len(relations) == 0 :
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '该学生不存在'}, json_dumps_params={'ensure_ascii': False})
    elif len(relations) >1 :
        return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '存在多个相同学号的学生'},json_dumps_params={'ensure_ascii': False})
    for item in stu_mes.keys():
        if item == "name":
            relations[0].xm = stu_mes.get(item)
        elif item == "xq":
            relations[0].xq = stu_mes.get(item)
        elif item == "sfzx":
            relations[0].sfzx = stu_mes.get(item)
        elif item == "cc":
            relations[0].cc = stu_mes.get(item)
        elif item == "glyx":
            relations[0].glyx = stu_mes.get(item)
        elif item == "instructor_name" or "instructor_num":
            relations2 = Instructor_Student.objects.filter(xh = stu_mes.get("xh"))
            if item == "instructor_name":
                if len(relations2)==0:
                    return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '该学生不存在'},json_dumps_params={'ensure_ascii': False})
                elif len(relations2)>1:
                    return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '存在多个相同学号的学生'},json_dumps_params={'ensure_ascii': False})
                else:
                    relations2[0].xm = stu_mes.get(item)
            elif item == "instructor_num":
                if len(relations2)==0:
                    return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '该学生不存在'},json_dumps_params={'ensure_ascii': False})
                elif len(relations2)>1:
                    return JsonResponse(status=HTTPStatus.NO_CONTENT, data={'error': '存在多个相同学号的学生'},json_dumps_params={'ensure_ascii': False})
                else:
                    relations2[0].zgh = stu_mes.get(item)
                    relations2[0].save()
    relations[0].save()
    return JsonResponse(data={'message': 'ok'}, json_dumps_params={'ensure_ascii': False})